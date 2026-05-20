from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from werkzeug.datastructures import FileStorage
from werkzeug.exceptions import BadRequest, NotFound

from app.extensions import db
from app.models import Customer, CustomerField, CustomerNote, ServiceOption, StaffMember, User
from app.services.service_catalog import is_services_enabled, list_active_services


@dataclass(frozen=True)
class ImportFieldDefinition:
    key: str
    label: str
    required: bool = False


@dataclass(frozen=True)
class ImportEntityDefinition:
    key: str
    label: str
    singular_label: str
    template_base_name: str
    fields: tuple[ImportFieldDefinition, ...]
    example_row: tuple[str, ...]


CUSTOMER_IMPORT = ImportEntityDefinition(
    key="customers",
    label="Customers",
    singular_label="customer",
    template_base_name="customer_import_template",
    fields=(
        ImportFieldDefinition("name", "Name", required=True),
        ImportFieldDefinition("phone", "Phone"),
        ImportFieldDefinition("email", "Email"),
        ImportFieldDefinition("city", "City", required=True),
        ImportFieldDefinition("notes", "Notes"),
        ImportFieldDefinition("billing_frequency", "Billing Frequency"),
        ImportFieldDefinition("billing_amount", "Billing Amount"),
    ),
    example_row=(
        "Jordan Harper",
        "555-333-2222",
        "jordan@example.com",
        "Springfield",
        "Imported from spring outreach list.",
        "monthly",
        "150.00",
    ),
)

STAFF_IMPORT = ImportEntityDefinition(
    key="staff",
    label="Employees / Staff",
    singular_label="staff member",
    template_base_name="employee_import_template",
    fields=(
        ImportFieldDefinition("name", "Name", required=True),
        ImportFieldDefinition("phone", "Phone"),
        ImportFieldDefinition("email", "Email"),
        ImportFieldDefinition("services", "Services"),
        ImportFieldDefinition("availability_notes", "Availability Notes"),
    ),
    example_row=(
        "Alex Crew",
        "555-111-2222",
        "alex@example.com",
        "Window Cleaning, Painting",
        "Prefers morning exterior work.",
    ),
)

STAFF_IMPORT_WITHOUT_SERVICES = ImportEntityDefinition(
    key="staff",
    label="Employees / Staff",
    singular_label="staff member",
    template_base_name="employee_import_template",
    fields=(
        ImportFieldDefinition("name", "Name", required=True),
        ImportFieldDefinition("phone", "Phone"),
        ImportFieldDefinition("email", "Email"),
        ImportFieldDefinition("availability_notes", "Availability Notes"),
    ),
    example_row=(
        "Alex Crew",
        "555-111-2222",
        "alex@example.com",
        "Prefers morning exterior work.",
    ),
)

IMPORT_ENTITIES = {
    CUSTOMER_IMPORT.key: CUSTOMER_IMPORT,
    STAFF_IMPORT.key: STAFF_IMPORT,
}

SERVICE_CELL_SPLIT_RE = re.compile(r"[,;\n]+")


def get_import_definition(entity_type: str) -> ImportEntityDefinition:
    normalized_entity_type = (entity_type or "").strip().lower()
    if normalized_entity_type == STAFF_IMPORT.key and not is_services_enabled():
        return STAFF_IMPORT_WITHOUT_SERVICES

    definition = IMPORT_ENTITIES.get(normalized_entity_type)
    if definition is None:
        raise NotFound("Import type not found.")
    return definition


def build_import_template(entity_type: str, file_format: str) -> tuple[bytes, str, str]:
    definition = get_import_definition(entity_type)
    normalized_format = (file_format or "").strip().lower()
    if normalized_format == "csv":
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow([field.key for field in definition.fields])
        writer.writerow(definition.example_row)
        return output.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", f"{definition.template_base_name}.csv"

    if normalized_format == "xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = _normalize_worksheet_title(definition.label)
        worksheet.append([field.key for field in definition.fields])
        worksheet.append(list(definition.example_row))
        for column_index, field in enumerate(definition.fields, start=1):
            worksheet.column_dimensions[_excel_column_name(column_index)].width = max(len(field.key), len(field.label), 18)

        output = io.BytesIO()
        workbook.save(output)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{definition.template_base_name}.xlsx",
        )

    raise BadRequest("Use CSV or XLSX templates only.")


def parse_import_upload(entity_type: str, uploaded_file: FileStorage) -> dict[str, Any]:
    definition = get_import_definition(entity_type)
    rows = _rows_from_uploaded_file(definition, uploaded_file)
    return preview_import_rows(entity_type, rows)


def preview_import_rows(entity_type: str, raw_rows: list[dict[str, Any]]) -> dict[str, Any]:
    definition = get_import_definition(entity_type)
    submitted_rows = _normalize_submitted_rows(definition, raw_rows)
    context = _build_preview_context(entity_type)
    preview_rows: list[dict[str, Any]] = []

    for row in submitted_rows:
        if entity_type == CUSTOMER_IMPORT.key:
            preview_rows.append(_preview_customer_row(row["rowNumber"], row["values"], context))
        elif entity_type == STAFF_IMPORT.key:
            preview_rows.append(_preview_staff_row(row["rowNumber"], row["values"], context))
        else:
            raise NotFound("Import type not found.")

    if not preview_rows:
        raise BadRequest("No data rows were found in the uploaded file.")

    return {
        "entityType": definition.key,
        "entityLabel": definition.label,
        "singularLabel": definition.singular_label,
        "fields": [
            {"key": field.key, "label": field.label, "required": field.required}
            for field in definition.fields
        ],
        "rows": preview_rows,
        "summary": {
            "totalRows": len(preview_rows),
            "readyRows": sum(1 for row in preview_rows if row["status"] == "ready"),
            "rowsWithIssues": sum(1 for row in preview_rows if row["issueCount"] > 0),
            "duplicateRows": sum(1 for row in preview_rows if row["duplicateCandidates"]),
            "errorRows": sum(1 for row in preview_rows if row["cellErrors"]),
        },
    }


def execute_import_rows(entity_type: str, submitted_rows: list[dict[str, Any]], user: User) -> dict[str, Any]:
    definition = get_import_definition(entity_type)
    normalized_rows = _normalize_submitted_rows(definition, submitted_rows)
    preview_payload = preview_import_rows(entity_type, normalized_rows)
    preview_rows_by_number = {row["rowNumber"]: row for row in preview_payload["rows"]}

    imported_count = 0
    merged_count = 0
    skipped_count = 0
    unresolved_issue_count = 0
    result_rows: list[dict[str, Any]] = []

    for raw_row in submitted_rows:
        row_number = _coerce_row_number(raw_row.get("rowNumber"))
        if row_number is None:
            continue

        preview_row = preview_rows_by_number.get(row_number)
        if preview_row is None:
            skipped_count += 1
            unresolved_issue_count += 1
            result_rows.append(
                {
                    "rowNumber": row_number,
                    "outcome": "unresolved",
                    "message": "This row could not be re-validated for import.",
                }
            )
            continue

        action = (raw_row.get("action") or preview_row.get("suggestedAction") or "").strip().lower()
        merge_target_id = _coerce_merge_target_id(raw_row.get("mergeTargetId"))

        if action == "skip":
            skipped_count += 1
            result_rows.append(
                {
                    "rowNumber": row_number,
                    "outcome": "skipped",
                    "message": "Skipped by user.",
                }
            )
            continue

        if preview_row["cellErrors"]:
            skipped_count += 1
            unresolved_issue_count += 1
            result_rows.append(
                {
                    "rowNumber": row_number,
                    "outcome": "unresolved",
                    "message": "This row still has validation issues.",
                }
            )
            continue

        if preview_row["duplicateCandidates"] and action not in {"create", "merge"}:
            skipped_count += 1
            unresolved_issue_count += 1
            result_rows.append(
                {
                    "rowNumber": row_number,
                    "outcome": "unresolved",
                    "message": "Choose whether to merge, create new, or skip this duplicate row.",
                }
            )
            continue

        if action == "merge":
            candidate_ids = {candidate["id"] for candidate in preview_row["duplicateCandidates"]}
            if merge_target_id is None or merge_target_id not in candidate_ids:
                skipped_count += 1
                unresolved_issue_count += 1
                result_rows.append(
                    {
                        "rowNumber": row_number,
                        "outcome": "unresolved",
                        "message": "Choose an existing record to merge into before importing.",
                    }
                )
                continue

        try:
            if entity_type == CUSTOMER_IMPORT.key:
                if action == "merge":
                    _merge_customer_from_import(preview_row["values"], merge_target_id, user)
                    merged_count += 1
                    result_rows.append(
                        {
                            "rowNumber": row_number,
                            "outcome": "merged",
                            "message": "Merged into an existing customer record.",
                        }
                    )
                else:
                    _create_customer_from_import(preview_row["values"], user)
                    imported_count += 1
                    result_rows.append(
                        {
                            "rowNumber": row_number,
                            "outcome": "imported",
                            "message": "Imported as a new customer.",
                        }
                    )
            elif entity_type == STAFF_IMPORT.key:
                if action == "merge":
                    _merge_staff_from_import(preview_row["values"], merge_target_id)
                    merged_count += 1
                    result_rows.append(
                        {
                            "rowNumber": row_number,
                            "outcome": "merged",
                            "message": "Merged into an existing staff record.",
                        }
                    )
                else:
                    _create_staff_from_import(preview_row["values"])
                    imported_count += 1
                    result_rows.append(
                        {
                            "rowNumber": row_number,
                            "outcome": "imported",
                            "message": "Imported as a new staff record.",
                        }
                    )
            else:
                raise NotFound("Import type not found.")
        except Exception as exc:
            db.session.rollback()
            skipped_count += 1
            unresolved_issue_count += 1
            result_rows.append(
                {
                    "rowNumber": row_number,
                    "outcome": "unresolved",
                    "message": str(exc),
                }
            )

    return {
        "entityType": definition.key,
        "entityLabel": definition.label,
        "summary": {
            "totalRows": len(preview_payload["rows"]),
            "importedCount": imported_count,
            "mergedCount": merged_count,
            "skippedCount": skipped_count,
            "unresolvedIssueCount": unresolved_issue_count,
        },
        "resultRows": result_rows,
    }


def _rows_from_uploaded_file(definition: ImportEntityDefinition, uploaded_file: FileStorage | None) -> list[dict[str, Any]]:
    if uploaded_file is None or not uploaded_file.filename:
        raise BadRequest("Choose a CSV or XLSX file to import.")

    filename = uploaded_file.filename.lower()
    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise BadRequest("The uploaded file is empty.")

    if filename.endswith(".csv"):
        rows = list(csv.reader(io.StringIO(file_bytes.decode("utf-8-sig"))))
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active
        rows = [[_coerce_cell_value(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    else:
        raise BadRequest("Use CSV or XLSX files only.")

    if not rows:
        raise BadRequest("The uploaded file is empty.")

    expected_headers = [field.key for field in definition.fields]
    actual_headers = [(_coerce_cell_value(value) or "").strip().lower() for value in rows[0]]
    if actual_headers != expected_headers:
        raise BadRequest(f"Use the provided template columns exactly: {', '.join(expected_headers)}.")

    parsed_rows = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = {}
        for index, field in enumerate(definition.fields):
            values[field.key] = _coerce_cell_value(row[index] if index < len(row) else "")
        if any(value for value in values.values()):
            parsed_rows.append({"rowNumber": row_number, "values": values})
    return parsed_rows


def _normalize_submitted_rows(definition: ImportEntityDefinition, submitted_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    for index, raw_row in enumerate(submitted_rows, start=2):
        raw_values = raw_row.get("values") if isinstance(raw_row, dict) else {}
        if not isinstance(raw_values, dict):
            raw_values = {}
        values = {field.key: (_coerce_cell_value(raw_values.get(field.key))).strip() for field in definition.fields}
        if not any(values.values()):
            continue
        normalized_rows.append(
            {
                "rowNumber": _coerce_row_number(raw_row.get("rowNumber")) or index,
                "values": values,
            }
        )
    return normalized_rows


def _build_preview_context(entity_type: str) -> dict[str, Any]:
    if entity_type == CUSTOMER_IMPORT.key:
        customers = list(
            db.session.scalars(
                select(Customer)
                .options(selectinload(Customer.fields))
                .order_by(Customer.primary_name, Customer.id)
            )
        )
        return {"customers": customers}

    if entity_type == STAFF_IMPORT.key:
        staff_members = list(db.session.scalars(select(StaffMember).order_by(StaffMember.display_name, StaffMember.id)))
        active_services = list_active_services() if is_services_enabled() else []
        return {
            "staffMembers": staff_members,
            "activeServicesByLowerName": {service.name.lower(): service for service in active_services},
        }

    raise NotFound("Import type not found.")


def _preview_customer_row(row_number: int, values: dict[str, str], context: dict[str, Any]) -> dict[str, Any]:
    normalized_values = {
        "name": values.get("name", "").strip(),
        "phone": values.get("phone", "").strip(),
        "email": values.get("email", "").strip().lower(),
        "city": values.get("city", "").strip(),
        "notes": values.get("notes", "").strip(),
        "billing_frequency": values.get("billing_frequency", "").strip().lower(),
        "billing_amount": values.get("billing_amount", "").strip(),
    }
    cell_errors: dict[str, str] = {}

    if not normalized_values["name"]:
        cell_errors["name"] = "Name is required."
    if not normalized_values["city"]:
        cell_errors["city"] = "City is required."

    if normalized_values["email"]:
        try:
            normalized_values["email"] = validate_email(
                normalized_values["email"],
                check_deliverability=False,
            ).normalized.lower()
        except EmailNotValidError:
            cell_errors["email"] = "Enter a valid email address."

    if normalized_values["phone"] and not _is_valid_phone(normalized_values["phone"]):
        cell_errors["phone"] = "Enter a valid phone number."

    if normalized_values["billing_frequency"] and normalized_values["billing_frequency"] not in Customer.BILLING_FREQUENCIES:
        cell_errors["billing_frequency"] = "Use weekly, monthly, or per_job."

    if normalized_values["billing_amount"]:
        try:
            billing_amount = Decimal(normalized_values["billing_amount"])
        except (InvalidOperation, ValueError):
            cell_errors["billing_amount"] = "Enter a valid billing amount."
        else:
            if billing_amount < 0:
                cell_errors["billing_amount"] = "Billing amount cannot be negative."

    duplicate_candidates = _find_customer_duplicate_candidates(normalized_values, context["customers"])
    issues = [
        {
            "code": "validation",
            "field": field_name,
            "message": message,
            "severity": "error",
        }
        for field_name, message in cell_errors.items()
    ]

    if duplicate_candidates:
        issues.append(
            {
                "code": "duplicate_existing",
                "field": None,
                "message": "Possible duplicate found. Choose whether to merge, create new, or skip this row.",
                "severity": "warning",
            }
        )

    return {
        "rowNumber": row_number,
        "values": normalized_values,
        "cellErrors": cell_errors,
        "issues": issues,
        "duplicateCandidates": duplicate_candidates,
        "issueCount": len(issues),
        "status": "ready" if not issues else "needs_review",
        "suggestedAction": "create" if not duplicate_candidates else "",
    }


def _preview_staff_row(row_number: int, values: dict[str, str], context: dict[str, Any]) -> dict[str, Any]:
    services_enabled = is_services_enabled()
    normalized_values = {
        "name": values.get("name", "").strip(),
        "phone": values.get("phone", "").strip(),
        "email": values.get("email", "").strip().lower(),
        "services": values.get("services", "").strip() if services_enabled else "",
        "availability_notes": values.get("availability_notes", "").strip(),
    }
    cell_errors: dict[str, str] = {}

    if not normalized_values["name"]:
        cell_errors["name"] = "Name is required."

    if normalized_values["email"]:
        try:
            normalized_values["email"] = validate_email(
                normalized_values["email"],
                check_deliverability=False,
            ).normalized.lower()
        except EmailNotValidError:
            cell_errors["email"] = "Enter a valid email address."

    if normalized_values["phone"] and not _is_valid_phone(normalized_values["phone"]):
        cell_errors["phone"] = "Enter a valid phone number."

    if services_enabled:
        service_names = _parse_service_names(normalized_values["services"])
        unknown_services = [
            service_name
            for service_name in service_names
            if service_name.lower() not in context["activeServicesByLowerName"]
        ]
        if unknown_services:
            cell_errors["services"] = f"Unknown services: {', '.join(unknown_services)}."

    duplicate_candidates = _find_staff_duplicate_candidates(normalized_values, context["staffMembers"])
    issues = [
        {
            "code": "validation",
            "field": field_name,
            "message": message,
            "severity": "error",
        }
        for field_name, message in cell_errors.items()
    ]

    if duplicate_candidates:
        issues.append(
            {
                "code": "duplicate_existing",
                "field": None,
                "message": "Possible duplicate found. Choose whether to merge, create new, or skip this row.",
                "severity": "warning",
            }
        )

    return {
        "rowNumber": row_number,
        "values": normalized_values,
        "cellErrors": cell_errors,
        "issues": issues,
        "duplicateCandidates": duplicate_candidates,
        "issueCount": len(issues),
        "status": "ready" if not issues else "needs_review",
        "suggestedAction": "create" if not duplicate_candidates else "",
    }


def _find_customer_duplicate_candidates(values: dict[str, str], customers: list[Customer]) -> list[dict[str, Any]]:
    email = values.get("email") or ""
    phone_digits = _normalize_phone_digits(values.get("phone") or "")
    name = (values.get("name") or "").strip().lower()
    city = (values.get("city") or "").strip().lower()

    candidates = []
    for customer in customers:
        matched_on = []
        email_values = {value.lower() for value in _customer_contact_values(customer, "email")}
        phone_values = {_normalize_phone_digits(value) for value in _customer_contact_values(customer, "phone") if value}

        if email and email in email_values:
            matched_on.append("email")
        if phone_digits and phone_digits in phone_values:
            matched_on.append("phone")
        if name and city and name == (customer.primary_name or "").strip().lower() and city == (customer.primary_city or "").strip().lower():
            matched_on.append("name and city")

        if matched_on:
            candidates.append(
                {
                    "id": customer.id,
                    "label": _format_customer_candidate(customer),
                    "matchedOn": matched_on,
                }
            )

    return candidates


def _find_staff_duplicate_candidates(values: dict[str, str], staff_members: list[StaffMember]) -> list[dict[str, Any]]:
    email = values.get("email") or ""
    phone_digits = _normalize_phone_digits(values.get("phone") or "")
    name = (values.get("name") or "").strip().lower()

    candidates = []
    for staff_member in staff_members:
        matched_on = []
        if email and email == (staff_member.email or "").strip().lower():
            matched_on.append("email")
        if phone_digits and phone_digits == _normalize_phone_digits(staff_member.phone or ""):
            matched_on.append("phone")
        if name and name == (staff_member.display_name or "").strip().lower():
            matched_on.append("name")

        if matched_on:
            candidates.append(
                {
                    "id": staff_member.id,
                    "label": _format_staff_candidate(staff_member),
                    "matchedOn": matched_on,
                }
            )

    return candidates


def _create_customer_from_import(values: dict[str, str], user: User) -> Customer:
    billing_amount = _parse_billing_amount(values.get("billing_amount") or "")
    customer = Customer(
        primary_name=values.get("name") or None,
        primary_phone=values.get("phone") or None,
        primary_email=values.get("email") or None,
        primary_city=values.get("city") or None,
        billing_amount=billing_amount,
        billing_frequency=(values.get("billing_frequency") or None),
    )
    db.session.add(customer)
    db.session.flush()
    _ensure_customer_field(customer, "name", customer.primary_name)
    _ensure_customer_field(customer, "phone", customer.primary_phone)
    _ensure_customer_field(customer, "email", customer.primary_email)
    _ensure_customer_field(customer, "city", customer.primary_city)

    if values.get("notes"):
        db.session.add(CustomerNote(customer_id=customer.id, note_text=values["notes"], created_by=user.id))

    db.session.commit()
    return customer


def _merge_customer_from_import(values: dict[str, str], customer_id: int | None, user: User) -> Customer:
    customer = db.session.get(Customer, customer_id)
    if customer is None:
        raise NotFound("Customer not found for merge.")

    if values.get("name"):
        customer.primary_name = values["name"]
    if values.get("phone"):
        customer.primary_phone = values["phone"]
    if values.get("email"):
        customer.primary_email = values["email"]
    if values.get("city"):
        customer.primary_city = values["city"]
    if values.get("billing_frequency"):
        customer.billing_frequency = values["billing_frequency"]
    if values.get("billing_amount"):
        customer.billing_amount = _parse_billing_amount(values["billing_amount"])

    _ensure_customer_field(customer, "name", customer.primary_name)
    _ensure_customer_field(customer, "phone", customer.primary_phone)
    _ensure_customer_field(customer, "email", customer.primary_email)
    _ensure_customer_field(customer, "city", customer.primary_city)

    if values.get("notes"):
        db.session.add(CustomerNote(customer_id=customer.id, note_text=values["notes"], created_by=user.id))

    db.session.commit()
    return customer


def _create_staff_from_import(values: dict[str, str]) -> StaffMember:
    service_objects = _resolve_staff_import_services(values.get("services") or "")
    staff_member = StaffMember(
        display_name=values.get("name") or "",
        phone=values.get("phone") or None,
        email=values.get("email") or None,
        worker_type="employee",
        status="active",
        notes=values.get("availability_notes") or None,
    )
    if service_objects:
        staff_member.services = service_objects
    db.session.add(staff_member)
    db.session.commit()
    return staff_member


def _merge_staff_from_import(values: dict[str, str], staff_member_id: int | None) -> StaffMember:
    staff_member = db.session.get(StaffMember, staff_member_id)
    if staff_member is None:
        raise NotFound("Staff member not found for merge.")

    if values.get("name"):
        staff_member.display_name = values["name"]
    if values.get("phone"):
        staff_member.phone = values["phone"]
    if values.get("email"):
        staff_member.email = values["email"]

    imported_services = _resolve_staff_import_services(values.get("services") or "")
    if imported_services:
        combined_services = {service.id: service for service in staff_member.services}
        for service in imported_services:
            combined_services[service.id] = service
        staff_member.services = sorted(
            combined_services.values(),
            key=lambda service: (service.display_order, service.name.lower(), service.id),
        )

    if values.get("availability_notes"):
        staff_member.notes = _merge_staff_notes(staff_member.notes, values["availability_notes"])

    db.session.commit()
    return staff_member


def _resolve_staff_import_services(raw_services: str) -> list[ServiceOption]:
    if not is_services_enabled():
        return []

    service_names = _parse_service_names(raw_services)
    if not service_names:
        return []

    active_services = {service.name.lower(): service for service in list_active_services()}
    resolved = []
    for service_name in service_names:
        service = active_services.get(service_name.lower())
        if service is None:
            raise BadRequest(f"Unknown service: {service_name}.")
        resolved.append(service)
    return resolved


def _merge_staff_notes(existing_notes: str | None, imported_note: str) -> str:
    imported_note = imported_note.strip()
    if not imported_note:
        return existing_notes or ""
    if not existing_notes:
        return imported_note
    if imported_note in existing_notes:
        return existing_notes
    return f"{existing_notes.strip()}\n\n{imported_note}"


def _normalize_worksheet_title(value: str) -> str:
    normalized = re.sub(r"[\\/*?:\[\]]", "-", value or "Sheet").strip()
    return (normalized or "Sheet")[:31]


def _ensure_customer_field(customer: Customer, kind: str, value: str | None) -> None:
    cleaned_value = (value or "").strip()
    if not cleaned_value:
        return

    existing = next(
        (
            field
            for field in customer.fields
            if field.kind == kind and (field.value or "").strip().lower() == cleaned_value.lower()
        ),
        None,
    )
    if existing is not None:
        existing.is_primary = True
        return

    db.session.add(
        CustomerField(
            customer_id=customer.id,
            kind=kind,
            value=cleaned_value,
            is_primary=True,
        )
    )


def _customer_contact_values(customer: Customer, kind: str) -> list[str]:
    values: list[str] = []
    primary_attr_by_kind = {
        "name": customer.primary_name,
        "phone": customer.primary_phone,
        "email": customer.primary_email,
        "city": customer.primary_city,
    }
    primary_value = primary_attr_by_kind.get(kind)
    if primary_value:
        values.append(primary_value)
    values.extend(field.value for field in customer.fields if field.kind == kind and field.value)
    return values


def _format_customer_candidate(customer: Customer) -> str:
    return (
        f"{customer.primary_name or 'Unnamed'} · "
        f"{customer.primary_email or 'no email'} · "
        f"{customer.primary_phone or 'no phone'}"
    )


def _format_staff_candidate(staff_member: StaffMember) -> str:
    return (
        f"{staff_member.display_name or 'Unnamed'} · "
        f"{staff_member.email or 'no email'} · "
        f"{staff_member.phone or 'no phone'}"
    )


def _parse_service_names(raw_services: str) -> list[str]:
    if not raw_services:
        return []
    service_names = []
    for part in SERVICE_CELL_SPLIT_RE.split(raw_services):
        cleaned = part.strip()
        if cleaned and cleaned not in service_names:
            service_names.append(cleaned)
    return service_names


def _parse_billing_amount(raw_amount: str) -> Decimal | None:
    if not raw_amount:
        return None
    try:
        billing_amount = Decimal(raw_amount)
    except (InvalidOperation, ValueError):
        raise BadRequest("Enter a valid billing amount.")
    if billing_amount < 0:
        raise BadRequest("Billing amount cannot be negative.")
    return billing_amount


def _is_valid_phone(raw_phone: str) -> bool:
    digit_count = len(_normalize_phone_digits(raw_phone))
    return 7 <= digit_count <= 15


def _normalize_phone_digits(raw_phone: str) -> str:
    return "".join(character for character in raw_phone if character.isdigit())


def _coerce_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_row_number(value: Any) -> int | None:
    try:
        row_number = int(value)
    except (TypeError, ValueError):
        return None
    return row_number if row_number > 0 else None


def _coerce_merge_target_id(value: Any) -> int | None:
    try:
        merge_target_id = int(value)
    except (TypeError, ValueError):
        return None
    return merge_target_id if merge_target_id > 0 else None


def _excel_column_name(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result